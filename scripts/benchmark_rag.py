"""RAG 召回率基准（500 条）：纯向量 vs 混合检索（向量 + 关键词 + RRF + 时效）.

语料：40 篇跨领域合成文档；查询：500 条（480 条有目标 + 20 条无目标负例）。
每条核心查询生成 3 个形式（原文 / 语气改写 / 同义改写），同时单独统计
160 条核心查询（较高难度子集）与全量 500 条的结果。
指标：recall@1 / recall@3 / recall@5 / MRR，按类别分组，负例统计误召回。
产出：逐条明细 + 汇总 + 分类 + 弱项 四张 Excel 工作表，保存到后端 docs 目录。

隔离性：基准语料使用独立 scene_tag="ragbench"，不触碰用户真实知识库；
结束后只清理本基准创建的文档与文件。
"""

import asyncio
import os
import sys
import time
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import bindparam, text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

from app.core.config import settings
from app.services.rag import create_space, process_document_pipeline, upload_document_file
from app.services.rag.embeddings import embed_query
from app.services.rag.knowledge import _scope_conditions, search_user_knowledge

USER_ID = "58b3f64f-0d22-4ef8-a79f-69c19e32b9b8"
SPACE_TAG = "ragbench"
TOPK = 5
EXCEL_DIR = Path(__file__).resolve().parent.parent / "docs"

# ── 语料：30 篇跨领域文档 ──────────────────────────────────
DOCS = [
    ("nginx_deploy.md", "deploy", "部署文档：生产环境使用 Nginx 反向代理，监听 80 端口，SSL 证书位于 /etc/nginx/ssl。负载均衡策略为加权轮询，后端节点共 4 台，权重 4:3:2:1。"),
    ("postgres_pool.md", "database", "数据库配置：PostgreSQL 连接池大小 20，最大溢出 10，空闲超时 300 秒。向量检索使用 pgvector，余弦距离，建立 HNSW 索引。"),
    ("redis_cache.md", "database", "缓存设计：Redis 缓存采用 Cache-Aside 模式，读未命中先查库再写缓存，Key 过期时间 300 秒。持久化同时开启 RDB 快照与 AOF 日志。"),
    ("react_state.md", "frontend", "前端开发：React 组件使用 useState 管理本地状态，useEffect 处理副作用，zustand 作为全局状态管理库，支持持久化中间件。"),
    ("vue_router.md", "frontend", "前端路由：Vue Router 使用 history 模式，路由组件懒加载，刷新页面时后端需将未知路径回退到 index.html，否则 404。"),
    ("game_boss.md", "game", "游戏攻略：暗影魔龙 BOSS 血量 500 万，第二阶段会释放全屏火焰，需要提前分散站位，坦克拉仇恨，治疗注意驱散，输出躲火圈。"),
    ("game_equip.md", "game", "游戏系统：装备强化上限 +10，强化失败不掉级，强化材料为陨铁，每次强化消耗金币与陨铁，+8 以上成功率逐步降低。"),
    ("meeting_q3.md", "meeting", "会议纪要：2026 年 8 月产品例会，讨论 Q3 预算分配，市场部 300 万，研发部 500 万，运营部 200 万，客服部 100 万，总预算 1100 万。"),
    ("history_tang.md", "history", "历史资料：唐朝贞观年间长安城人口超百万，东西两市商业繁荣，丝绸之路贸易往来频繁，胡商云集。"),
    ("history_ming.md", "history", "历史资料：明朝永乐年间郑和七下西洋，1405 年首次出发，宝船六十余艘，最远到达非洲东海岸。"),
    ("news_lumi2.md", "news", "新闻：2026 年 8 月 9 日，公司发布新一代 AI 助手 Lumi 2.0，支持多模态对话、本地知识库与多智能体协作办公。"),
    ("news_policy.md", "news", "新闻：国家发布《人工智能生成内容管理办法》，要求深度合成内容显式标注，2026 年 9 月 1 日起施行。"),
    ("pagination.py", "code", "def paginate(items, page, size):\n    start = (page - 1) * size\n    return items[start:start + size]\n\ndef total_pages(count, size):\n    import math\n    return math.ceil(count / size)"),
    ("auth_jwt.py", "code", "def create_token(user_id):\n    payload = {\"sub\": user_id, \"exp\": now + 7 * 86400}\n    return jwt.encode(payload, SECRET, algorithm=\"HS256\")\n\ndef verify_token(token):\n    return jwt.decode(token, SECRET, algorithms=[\"HS256\"])"),
    ("budget_table.md", "table", "# 2026 Q3 预算表\n\n| 部门 | 预算(万) |\n| --- | --- |\n| 市场部 | 300 |\n| 研发部 | 500 |\n| 运营部 | 200 |\n| 客服部 | 100 |"),
    ("sales_table.md", "table", "# 2025 各季度销售额(万元)\n\n| 季度 | 销售额 |\n| --- | --- |\n| Q1 | 120 |\n| Q2 | 150 |\n| Q3 | 180 |\n| Q4 | 200 |"),
    ("ops_runbook.md", "ops", "运维手册：密钥轮换任务编号 ROTATE-7721，每周日凌晨三点执行，执行窗口 30 分钟，失败自动重试 3 次，重试间隔 5 分钟。"),
    ("monitor_alert.md", "ops", "监控告警：CPU 使用率超过 90% 持续 5 分钟触发告警，通过钉钉机器人推送，同时短信通知值班电话，夜间静默时段不推送短信。"),
    ("medical_diet.md", "health", "健康指南：高血压患者每日食盐摄入不超过 5 克，多吃富含钾的蔬菜水果，少吃腌制食品，定期监测血压。"),
    ("fitness_plan.md", "health", "健身计划：减脂期每周进行 4 次有氧运动，每次 40 分钟，蛋白质摄入每公斤体重 1.6 克，力量训练穿插进行。"),
    ("finance_fund.md", "finance", "理财知识：指数基金定投适合长期投资，历史年化收益 8%-10%，申购费一般 0.15%，定投可以摊平成本。"),
    ("finance_tax.md", "finance", "税务政策：个税综合所得起征点每月 5000 元，超额累进税率 3%-45%，专项附加扣除包括子女教育、住房贷款利息等。"),
    ("edu_exam.md", "edu", "教育信息：全国统一高考每年 6 月 7 日开始，考试科目为 3+X，语文数学外语为统考科目，X 由各省自定。"),
    ("edu_course.md", "edu", "课程介绍：Python 入门课程共 8 周，每周 1 次线下课加 2 次线上练习，结课需要完成一个数据分析小项目。"),
    ("law_contract.md", "law", "法律常识：劳动合同试用期最长不超过 6 个月，解除劳动合同需提前 30 天书面通知，试用期工资不得低于转正的 80%。"),
    ("law_privacy.md", "law", "法律常识：个人信息保护法要求收集个人信息遵循最小必要原则，处理敏感个人信息需取得单独同意。"),
    ("travel_yunnan.md", "travel", "旅游攻略：云南大理洱海环湖骑行约 120 公里，丽江古城夜晚热闹，玉龙雪山海拔 5596 米，需提前预约索道。"),
    ("food_hotpot.md", "food", "美食攻略：重庆火锅用九宫格锅底，毛肚涮 7 秒口感最佳，鸭肠涮 10 秒，脑花需要煮 15 分钟。"),
    ("car_ev.md", "car", "汽车资讯：新款纯电 SUV 综合续航 600 公里，快充 30 分钟可从 20% 充到 80%，搭载 L2 级辅助驾驶。"),
    ("sport_marathon.md", "sport", "体育知识：全程马拉松距离 42.195 公里，关门时间一般为 6 小时，补给站每 5 公里一个，配速 5 分半可完赛。"),
    ("astronomy_solar.md", "science", "天文知识：太阳系有八大行星，按距太阳远近依次为水星、金星、地球、火星、木星、土星、天王星、海王星，木星体积最大。"),
    ("chem_water.md", "science", "化学知识：水的化学式是 H2O，由两个氢原子和一个氧原子构成，标准大气压下沸点 100 摄氏度，冰点 0 摄氏度。"),
    ("bio_cell.md", "science", "生物知识：细胞是生命活动的基本单位，动物细胞含细胞膜、细胞质、细胞核，植物细胞额外有细胞壁和叶绿体。"),
    ("geo_himalaya.md", "geography", "地理知识：珠穆朗玛峰海拔 8848.86 米，位于中国与尼泊尔边境，属于喜马拉雅山脉，是世界最高峰。"),
    ("weather_typhoon.md", "weather", "气象知识：台风中心气压越低强度越强，风力达到 12 级及以上称为强台风，登陆前需提前 24 小时发布预警。"),
    ("music_piano.md", "art", "音乐知识：钢琴有 88 个琴键，白键 52 个黑键 36 个，标准音 A4 频率 440 赫兹，踏板分延音、弱音、持音三种。"),
    ("movie_sci.md", "art", "影视资讯：科幻电影《星际穿越》由诺兰执导，讲述穿越虫洞寻找新家园的故事，黑洞视觉特效由真实物理计算生成。"),
    ("linux_ops.md", "tech", "操作系统：Linux 常用命令 ls 列出文件，grep 搜索文本，chmod 修改权限，ps 查看进程，df 查看磁盘使用。"),
    ("network_tcp.md", "tech", "网络协议：TCP 三次握手建立连接，四次挥手断开连接，SYN 泛洪攻击利用半连接耗尽服务器资源。"),
    ("ai_llm.md", "tech", "人工智能：大语言模型基于 Transformer 架构，通过自注意力机制处理文本，RLHF 人类反馈强化学习提升对齐能力。"),
]

# ── 核心查询：160 条有目标（40 篇 × 4 条，含较高难度） ──────
CORE_QUERIES = [
    # nginx_deploy
    ("Nginx 反向代理怎么配置", "nginx_deploy.md"),
    ("网站负载均衡怎么做", "nginx_deploy.md"),
    ("SSL 证书放在服务器哪个目录", "nginx_deploy.md"),
    ("后端有几台服务器，负载权重怎么分配的", "nginx_deploy.md"),
    # postgres_pool
    ("PostgreSQL 连接池大小设置多少合适", "postgres_pool.md"),
    ("pgvector 用什么距离做向量相似检索", "postgres_pool.md"),
    ("数据库连接空闲多久超时", "postgres_pool.md"),
    ("向量检索建了什么索引", "postgres_pool.md"),
    # redis_cache
    ("Redis 缓存读不到数据怎么处理", "redis_cache.md"),
    ("缓存 Key 过期时间设多长", "redis_cache.md"),
    ("Redis 持久化用 RDB 还是 AOF", "redis_cache.md"),
    ("Cache-Aside 缓存模式是什么意思", "redis_cache.md"),
    # react_state
    ("React 怎么管理全局状态", "react_state.md"),
    ("useState 和 useEffect 分别是干什么的", "react_state.md"),
    ("zustand 支持持久化吗", "react_state.md"),
    ("React 组件本地状态用什么钩子", "react_state.md"),
    # vue_router
    ("Vue Router history 模式刷新页面 404 怎么办", "vue_router.md"),
    ("路由组件懒加载怎么配置", "vue_router.md"),
    ("前端单页路由怎么实现", "vue_router.md"),
    ("未知路径怎么回退到首页", "vue_router.md"),
    # game_boss
    ("暗影魔龙第二阶段怎么打", "game_boss.md"),
    ("BOSS 全屏火焰技能怎么躲", "game_boss.md"),
    ("暗影魔龙血量有多少", "game_boss.md"),
    ("打暗影魔龙治疗职业要注意什么", "game_boss.md"),
    # game_equip
    ("装备强化失败会掉级吗", "game_equip.md"),
    ("强化装备需要什么材料", "game_equip.md"),
    ("装备强化上限是多少", "game_equip.md"),
    ("装备 +8 以后强化成功率怎么样", "game_equip.md"),
    # meeting_q3
    ("Q3 预算怎么分配的", "meeting_q3.md"),
    ("市场部研发部运营部各分多少钱", "meeting_q3.md"),
    ("这次产品例会的总预算是多少", "meeting_q3.md"),
    ("客服部预算额度是多少", "meeting_q3.md"),
    # history_tang
    ("唐朝长安城有多少人口", "history_tang.md"),
    ("长安城东西两市是干什么的", "history_tang.md"),
    ("丝绸之路在哪个朝代最繁荣", "history_tang.md"),
    ("贞观年间长安商业怎么样", "history_tang.md"),
    # history_ming
    ("郑和第一次下西洋是哪一年", "history_ming.md"),
    ("郑和宝船有多少艘", "history_ming.md"),
    ("郑和最远航行到过哪里", "history_ming.md"),
    ("郑和一共下过几次西洋", "history_ming.md"),
    # news_lumi2
    ("Lumi 2.0 什么时候发布的", "news_lumi2.md"),
    ("Lumi 2.0 支持哪些新功能", "news_lumi2.md"),
    ("公司最近发布了什么新产品", "news_lumi2.md"),
    ("多智能体协作办公是什么", "news_lumi2.md"),
    # news_policy
    ("AI 生成内容管理办法什么时候施行", "news_policy.md"),
    ("深度合成内容需要标注吗", "news_policy.md"),
    ("国家出台了什么 AI 新规", "news_policy.md"),
    ("AI 内容管理新规要求什么", "news_policy.md"),
    # pagination
    ("分页函数怎么写", "pagination.py"),
    ("怎么计算总页数", "pagination.py"),
    ("分页时起始位置怎么算", "pagination.py"),
    ("math.ceil 怎么算页数", "pagination.py"),
    # auth_jwt
    ("JWT token 怎么生成", "auth_jwt.py"),
    ("JWT 有效期一般设多长", "auth_jwt.py"),
    ("怎么校验 token 是否合法", "auth_jwt.py"),
    ("HS256 签名算法怎么用", "auth_jwt.py"),
    # budget_table
    ("各部门预算额度是多少", "budget_table.md"),
    ("市场部预算多少万", "budget_table.md"),
    ("预算表里运营部是多少", "budget_table.md"),
    ("哪个部门预算最高", "budget_table.md"),
    # sales_table
    ("2025 年 Q2 销售额是多少", "sales_table.md"),
    ("各季度销售额分别是多少", "sales_table.md"),
    ("去年第四季度卖了多少", "sales_table.md"),
    ("销售额最高的季度是哪个", "sales_table.md"),
    # ops_runbook
    ("ROTATE-7721 是什么任务", "ops_runbook.md"),
    ("密钥轮换任务几点执行", "ops_runbook.md"),
    ("密钥轮换失败会重试几次", "ops_runbook.md"),
    ("轮换任务执行窗口有多久", "ops_runbook.md"),
    # monitor_alert
    ("CPU 使用率超过多少会告警", "monitor_alert.md"),
    ("监控告警通知怎么发出去", "monitor_alert.md"),
    ("短信通知值班电话什么时候发", "monitor_alert.md"),
    ("监控告警的触发阈值是多少", "monitor_alert.md"),
    # medical_diet
    ("高血压患者一天吃多少盐", "medical_diet.md"),
    ("高血压适合多吃什么", "medical_diet.md"),
    ("高血压能吃腌制食品吗", "medical_diet.md"),
    ("高血压饮食要注意什么", "medical_diet.md"),
    # fitness_plan
    ("减脂期每周运动几次", "fitness_plan.md"),
    ("减脂期蛋白质要摄入多少", "fitness_plan.md"),
    ("每次有氧运动做多久", "fitness_plan.md"),
    ("减脂期力量训练怎么安排", "fitness_plan.md"),
    # finance_fund
    ("指数基金定投收益怎么样", "finance_fund.md"),
    ("基金申购费率是多少", "finance_fund.md"),
    ("基金定投有什么好处", "finance_fund.md"),
    ("长期投资适合选什么", "finance_fund.md"),
    # finance_tax
    ("个税起征点是多少", "finance_tax.md"),
    ("个税税率最高是多少", "finance_tax.md"),
    ("个税专项附加扣除有哪些", "finance_tax.md"),
    ("工资个税综合所得怎么算", "finance_tax.md"),
    # edu_exam
    ("高考每年什么时候开始", "edu_exam.md"),
    ("高考统考考哪些科目", "edu_exam.md"),
    ("高考 3+X 是什么意思", "edu_exam.md"),
    ("高考 X 科目由谁决定", "edu_exam.md"),
    # edu_course
    ("Python 入门课程一共几周", "edu_course.md"),
    ("Python 课程每周上几次", "edu_course.md"),
    ("Python 课结课要做什么", "edu_course.md"),
    ("这个 Python 课程怎么安排", "edu_course.md"),
    # law_contract
    ("劳动合同试用期最长多久", "law_contract.md"),
    ("辞职需要提前多久通知", "law_contract.md"),
    ("试用期工资是转正的多少", "law_contract.md"),
    ("解除劳动合同有什么要求", "law_contract.md"),
    # law_privacy
    ("收集个人信息要遵循什么原则", "law_privacy.md"),
    ("处理敏感个人信息需要什么", "law_privacy.md"),
    ("个人信息保护法核心要求是什么", "law_privacy.md"),
    ("最小必要原则是什么意思", "law_privacy.md"),
    # travel_yunnan
    ("大理洱海环湖骑行多少公里", "travel_yunnan.md"),
    ("玉龙雪山海拔多高", "travel_yunnan.md"),
    ("丽江古城晚上怎么样", "travel_yunnan.md"),
    ("去玉龙雪山要提前预约吗", "travel_yunnan.md"),
    # food_hotpot
    ("毛肚涮几秒最好吃", "food_hotpot.md"),
    ("重庆火锅锅底是什么样", "food_hotpot.md"),
    ("鸭肠要涮多久", "food_hotpot.md"),
    ("脑花要煮多久", "food_hotpot.md"),
    # car_ev
    ("这款电车续航多少公里", "car_ev.md"),
    ("快充多久能充到 80%", "car_ev.md"),
    ("L2 辅助驾驶是什么等级", "car_ev.md"),
    ("电车从 20% 充到 80% 要多久", "car_ev.md"),
    # sport_marathon
    ("全程马拉松多少公里", "sport_marathon.md"),
    ("马拉松关门时间一般多久", "sport_marathon.md"),
    ("马拉松补给站隔多远一个", "sport_marathon.md"),
    ("5 分半配速能完赛吗", "sport_marathon.md"),
    # astronomy_solar
    ("太阳系有哪几大行星", "astronomy_solar.md"),
    ("按离太阳距离排序的行星顺序是什么", "astronomy_solar.md"),
    ("太阳系体积最大的行星是哪个", "astronomy_solar.md"),
    ("八大行星从近到远怎么排", "astronomy_solar.md"),
    # chem_water
    ("水的化学式是什么", "chem_water.md"),
    ("水分子由什么组成", "chem_water.md"),
    ("标准大气压下水的沸点是多少", "chem_water.md"),
    ("水的冰点是多少度", "chem_water.md"),
    # bio_cell
    ("动物细胞有哪些结构", "bio_cell.md"),
    ("植物细胞和动物细胞有什么不同", "bio_cell.md"),
    ("细胞是生命活动的基本单位吗", "bio_cell.md"),
    ("叶绿体在哪种细胞里有", "bio_cell.md"),
    # geo_himalaya
    ("珠穆朗玛峰海拔多少米", "geo_himalaya.md"),
    ("世界最高峰是哪个", "geo_himalaya.md"),
    ("珠峰位于哪个山脉", "geo_himalaya.md"),
    ("珠峰在中国和哪个国家的边境", "geo_himalaya.md"),
    # weather_typhoon
    ("台风强度怎么判断", "weather_typhoon.md"),
    ("多少级风算强台风", "weather_typhoon.md"),
    ("台风登陆前提前多久发布预警", "weather_typhoon.md"),
    ("台风中心气压低说明什么", "weather_typhoon.md"),
    # music_piano
    ("钢琴有多少个琴键", "music_piano.md"),
    ("标准音 A4 的频率是多少", "music_piano.md"),
    ("钢琴踏板有哪几种", "music_piano.md"),
    ("钢琴白键黑键各有多少个", "music_piano.md"),
    # movie_sci
    ("星际穿越讲的是什么故事", "movie_sci.md"),
    ("星际穿越的导演是谁", "movie_sci.md"),
    ("星际穿越的黑洞特效是怎么做的", "movie_sci.md"),
    ("哪部电影讲穿越虫洞找新家园", "movie_sci.md"),
    # linux_ops
    ("Linux 查看进程用什么命令", "linux_ops.md"),
    ("Linux 修改文件权限的命令是什么", "linux_ops.md"),
    ("Linux 怎么搜索文本内容", "linux_ops.md"),
    ("df 命令是干什么的", "linux_ops.md"),
    # network_tcp
    ("TCP 三次握手是什么", "network_tcp.md"),
    ("TCP 连接怎么断开", "network_tcp.md"),
    ("SYN 泛洪攻击是什么", "network_tcp.md"),
    ("TCP 建立连接的过程", "network_tcp.md"),
    # ai_llm
    ("大语言模型基于什么架构", "ai_llm.md"),
    ("RLHF 是什么", "ai_llm.md"),
    ("自注意力机制是干什么的", "ai_llm.md"),
    ("大模型怎么对齐人类偏好", "ai_llm.md"),
]

# ── 负例（无目标文档，应未命中；20 条） ────────────────────
NEGATIVES = [
    "量子计算的纠错码有哪些种类",
    "火星探测器祝融号发现了什么",
    "docker compose 如何配置卷挂载",
    "K8s 的 pod 生命周期钩子有哪些",
    "比特币白皮书的作者是谁",
    "贝多芬第五交响曲有几个乐章",
    "大熊猫的饮食习惯是什么",
    "2024 年巴黎奥运会金牌榜",
    "屈原的《离骚》一共多少字",
    "埃及金字塔是什么时候建造的",
    "Wi-Fi 7 支持哪些频段",
    "无人机驾驶证怎么考",
    "茅台酒的酿造工艺有哪些步骤",
    "猫咪绝育后要注意什么",
    "比特币挖矿一年耗电多少",
    "北欧神话的主神是谁",
    "乒乓球正式比赛每局几分",
    "电动汽车电池回收有什么政策",
    "碳中和具体是什么意思",
    "春节联欢晚会的节目单在哪看",
]


def _variants(q: str):
    """每条核心查询生成 3 个形式：原文 / 语气改写 / 同义改写."""
    out = [(q, "核心")]
    out.append(("我想知道" + q, "变体-语气"))
    alt = q.replace("怎么", "如何").replace("是什么", "指的是什么")
    if alt != q:
        out.append((alt, "变体-同义"))
    else:
        out.append(("请问，" + q + "？", "变体-语气"))
    return out


QUERIES: list[tuple[str, str | None, str]] = []
for _q, _t in CORE_QUERIES:
    for _text, _qtype in _variants(_q):
        QUERIES.append((_text, _t, _qtype))
for _neg in NEGATIVES:
    QUERIES.append((_neg, None, "负例"))


def _rank_of(results, target):
    """目标文档在结果中的位置（1 起始）；未命中返回 None."""
    for i, r in enumerate(results[:TOPK], 1):
        if r.get("title") == target:
            return i
    return None


async def vector_only(session, query, space_id):
    """纯向量检索（阈值过滤 + 空间/标签隔离），复现旧行为."""
    vec = await embed_query(query)
    if not vec:
        return []
    conds, params = _scope_conditions(USER_ID, [SPACE_TAG], need_embedding=True)
    sql = f"""
        SELECT d.filename AS title, 1 - (c.embedding <=> CAST(:qvec AS vector)) AS similarity
        FROM document_chunks c
        JOIN documents d ON d.id = c.document_id
        JOIN knowledge_spaces s ON s.id = c.space_id
        WHERE {' AND '.join(conds)} AND s.id = :space_id
        ORDER BY c.embedding <=> CAST(:qvec AS vector)
        LIMIT :top_k
    """
    params["space_id"] = space_id
    stmt = text(sql).bindparams(
        bindparam("qvec", "[" + ",".join(repr(float(x)) for x in vec) + "]"),
        bindparam("top_k", TOPK),
    )
    if "tags" in params:
        stmt = stmt.bindparams(bindparam("tags", expanding=True))
    rows = (await session.execute(stmt, params)).mappings().all()
    return [
        {"title": r["title"], "similarity": float(r["similarity"])}
        for r in rows
        if float(r["similarity"]) >= settings.RAG_SIMILARITY_THRESHOLD
    ]


def _fmt_rank(r):
    return "未命中" if r is None else str(r)


def _fmt_hit(r):
    return "✓" if r is not None else "✗"


def _summary_stats(ranks, total):
    """由 rank 列表计算 recall@1/3/5 与 MRR."""
    n = len(ranks)
    hits = {1: 0, 3: 0, 5: 0}
    mrr = 0.0
    miss = 0
    for r in ranks:
        if r is None:
            miss += 1
            continue
        mrr += 1.0 / r
        for k in (1, 3, 5):
            if r <= k:
                hits[k] += 1
    return {
        "total": total,
        "hit1": hits[1],
        "hit3": hits[3],
        "hit5": hits[5],
        "miss": miss,
        "mrr": mrr / total if total else 0.0,
    }


def _style_sheet(ws, headers, widths):
    head_fill = PatternFill("solid", fgColor="4472C4")
    head_font = Font(color="FFFFFF", bold=True)
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = "A2"


async def main() -> None:
    engine = create_async_engine(settings.DATABASE_URL, poolclass=NullPool)
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    space_id = None
    created_files: list[str] = []
    try:
        # 1) 建空间 + 上传语料
        async with factory() as session:
            space = await create_space(session, USER_ID, "RagBenchFull", "rag 召回率基准", SPACE_TAG)
            space_id = str(space.id)
            paths = {}
            for fname, _cat, content in DOCS:
                doc, path, _ = await upload_document_file(
                    session, USER_ID, space_id, fname, content.encode("utf-8")
                )
                paths[str(doc.id)] = (fname, str(path))
            await session.commit()

        # 2) 解析 + 清洗 + 分块 + 嵌入
        async with factory() as session:
            for doc_id, (fname, path) in paths.items():
                await process_document_pipeline(session, doc_id, path)
                created_files.append(path)
        print(f"[1/4] 语料入库完成: {len(paths)} 篇，开始检索评测…\n")

        # 3) 逐条评测
        detail_rows = []
        weak_rows = []
        cat_stats: dict[str, dict] = {}
        neg_rows = []
        vec_ranks, hyb_ranks = [], []
        core_vec_ranks, core_hyb_ranks = [], []
        target_queries = [(q, t, qt) for q, t, qt in QUERIES if t]
        neg_queries = [(q, t, qt) for q, t, qt in QUERIES if not t]

        async with factory() as session:
            for idx, (q, target, qtype) in enumerate(target_queries, 1):
                cat = next((c for fn, c, _ in DOCS if fn == target), "other")
                v = await vector_only(session, q, space_id)
                _, h = await search_user_knowledge(
                    session, USER_ID, q, [SPACE_TAG], top_k=TOPK
                )
                rv = _rank_of(v, target)
                rh = _rank_of(h, target)
                vec_ranks.append(rv)
                hyb_ranks.append(rh)
                if qtype == "核心":
                    core_vec_ranks.append(rv)
                    core_hyb_ranks.append(rh)
                cat_stats.setdefault(cat, {"vec": [], "hyb": []})
                cat_stats[cat]["vec"].append(rv)
                cat_stats[cat]["hyb"].append(rh)

                if rh is None and rv is None:
                    note = "双路都未命中"
                    weak_rows.append((q, target, _fmt_rank(rv), _fmt_rank(rh), "双路都未命中"))
                elif rv is None and rh is not None:
                    note = "混合救回"
                    weak_rows.append((q, target, _fmt_rank(rv), _fmt_rank(rh), "混合救回"))
                elif rv is not None and rh is None:
                    note = "混合反而丢失!"
                    weak_rows.append((q, target, _fmt_rank(rv), _fmt_rank(rh), "混合反而丢失"))
                else:
                    note = "双路命中"
                detail_rows.append(
                    (idx, q, target, cat, qtype, _fmt_rank(rv), _fmt_rank(rh),
                     _fmt_hit(rv), _fmt_hit(rh), note)
                )
                if idx % 20 == 0 or idx == len(target_queries):
                    print(f"  进度 {idx}/{len(target_queries)}…")

            # 负例：统计误召回
            for q, _t, _qt in neg_queries:
                v = await vector_only(session, q, space_id)
                _, h = await search_user_knowledge(
                    session, USER_ID, q, [SPACE_TAG], top_k=TOPK
                )
                v_hit = bool(v)
                h_hit = bool(h)
                neg_rows.append(
                    (q, "向量:" + ("误召回" if v_hit else "未召回"),
                     "混合:" + ("误召回" if h_hit else "未召回"))
                )

        n = len(target_queries)
        vec_s = _summary_stats(vec_ranks, n)
        hyb_s = _summary_stats(hyb_ranks, n)
        core_n = len(core_vec_ranks)
        core_vec_s = _summary_stats(core_vec_ranks, core_n)
        core_hyb_s = _summary_stats(core_hyb_ranks, core_n)
        neg_v = sum(1 for r in neg_rows if "误召回" in r[1])
        neg_h = sum(1 for r in neg_rows if "误召回" in r[2])

        # 4) 导出 Excel
        EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = EXCEL_DIR / f"rag_benchmark_{ts}.xlsx"

        wb = Workbook()
        # ── 明细 ──
        ws = wb.active
        ws.title = "逐条明细"
        headers = ["序号", "查询", "目标文档", "类别", "查询类型", "纯向量rank", "混合rank", "纯向量@5", "混合@5", "备注"]
        _style_sheet(ws, headers, [7, 46, 20, 9, 11, 11, 11, 11, 11, 18])
        for row in detail_rows:
            ws.append(row)
        # ── 汇总 ──
        ws = wb.create_sheet("汇总")
        _style_sheet(ws, ["指标", "纯向量", "混合检索", "差值(混合-向量)"], [16, 14, 14, 16])
        summary = [
            ("总查询数", n, n, ""),
            ("recall@1", vec_s["hit1"], hyb_s["hit1"], hyb_s["hit1"] - vec_s["hit1"]),
            ("recall@3", vec_s["hit3"], hyb_s["hit3"], hyb_s["hit3"] - vec_s["hit3"]),
            ("recall@5", vec_s["hit5"], hyb_s["hit5"], hyb_s["hit5"] - vec_s["hit5"]),
            ("recall@1 占比", f"{vec_s['hit1']}/{n} ({vec_s['hit1']/n:.1%})", f"{hyb_s['hit1']}/{n} ({hyb_s['hit1']/n:.1%})", ""),
            ("recall@3 占比", f"{vec_s['hit3']}/{n} ({vec_s['hit3']/n:.1%})", f"{hyb_s['hit3']}/{n} ({hyb_s['hit3']/n:.1%})", ""),
            ("recall@5 占比", f"{vec_s['hit5']}/{n} ({vec_s['hit5']/n:.1%})", f"{hyb_s['hit5']}/{n} ({hyb_s['hit5']/n:.1%})", ""),
            ("MRR", f"{vec_s['mrr']:.4f}", f"{hyb_s['mrr']:.4f}", f"{hyb_s['mrr'] - vec_s['mrr']:+.4f}"),
            ("未命中数", vec_s["miss"], hyb_s["miss"], ""),
            ("负例误召回", f"{neg_v}/{len(neg_queries)}", f"{neg_h}/{len(neg_queries)}", ""),
        ]
        for row in summary:
            ws.append(row)
        ws.append(("", "", "", ""))
        ws.append(("【核心查询子集（较高难度）】", "", "", ""))
        core_summary = [
            ("核心查询数", core_n, core_n, ""),
            ("recall@1", f"{core_vec_s['hit1']}/{core_n} ({core_vec_s['hit1']/core_n:.1%})", f"{core_hyb_s['hit1']}/{core_n} ({core_hyb_s['hit1']/core_n:.1%})", ""),
            ("recall@3", f"{core_vec_s['hit3']}/{core_n} ({core_vec_s['hit3']/core_n:.1%})", f"{core_hyb_s['hit3']}/{core_n} ({core_hyb_s['hit3']/core_n:.1%})", ""),
            ("recall@5", f"{core_vec_s['hit5']}/{core_n} ({core_vec_s['hit5']/core_n:.1%})", f"{core_hyb_s['hit5']}/{core_n} ({core_hyb_s['hit5']/core_n:.1%})", ""),
            ("MRR", f"{core_vec_s['mrr']:.4f}", f"{core_hyb_s['mrr']:.4f}", ""),
            ("未命中数", core_vec_s["miss"], core_hyb_s["miss"], ""),
        ]
        for row in core_summary:
            ws.append(row)
        # ── 分类 ──
        ws = wb.create_sheet("按类别")
        _style_sheet(ws, ["类别", "查询数", "向量@1", "向量@3", "向量@5", "混合@1", "混合@3", "混合@5"], [12, 9, 10, 10, 10, 10, 10, 10])
        for cat, d in sorted(cat_stats.items()):
            vs = _summary_stats(d["vec"], len(d["vec"]))
            hs = _summary_stats(d["hyb"], len(d["hyb"]))
            ws.append((cat, len(d["vec"]), vs["hit1"], vs["hit3"], vs["hit5"], hs["hit1"], hs["hit3"], hs["hit5"]))
        # ── 弱项 ──
        ws = wb.create_sheet("弱项与差异")
        _style_sheet(ws, ["查询", "目标文档", "纯向量rank", "混合rank", "说明"], [46, 20, 11, 11, 18])
        for row in weak_rows:
            ws.append(row)
        # ── 负例 ──
        ws = wb.create_sheet("负例")
        _style_sheet(ws, ["查询", "纯向量", "混合检索"], [46, 14, 14])
        for row in neg_rows:
            ws.append(row)

        wb.save(xlsx_path)
        print(f"\n[4/4] 已导出: {xlsx_path}")

        # 控制台汇总
        print(f"\n共 {n} 条有目标查询 + {len(neg_queries)} 条负例（全量 {n + len(neg_queries)} 条）, top_k={TOPK}")
        for label, s in (("纯向量", vec_s), ("混合", hyb_s)):
            print(f"{label}: recall@1={s['hit1']}/{n} ({s['hit1']/n:.1%}) | "
                  f"@3={s['hit3']}/{n} ({s['hit3']/n:.1%}) | "
                  f"@5={s['hit5']}/{n} ({s['hit5']/n:.1%}) | MRR={s['mrr']:.4f} | miss={s['miss']}")
        print(f"\n核心查询子集（{core_n} 条）:")
        for label, s in (("纯向量", core_vec_s), ("混合", core_hyb_s)):
            print(f"{label}: recall@1={s['hit1']}/{core_n} ({s['hit1']/core_n:.1%}) | "
                  f"@3={s['hit3']}/{core_n} ({s['hit3']/core_n:.1%}) | "
                  f"@5={s['hit5']}/{core_n} ({s['hit5']/core_n:.1%}) | MRR={s['mrr']:.4f} | miss={s['miss']}")
        print(f"负例误召回: 向量 {neg_v}/{len(neg_queries)} | 混合 {neg_h}/{len(neg_queries)}")
        if weak_rows:
            print("\n弱项/差异:")
            for row in weak_rows:
                print(f"  - {row[0][:30]} | {row[1]} | vec={row[2]} | hyb={row[3]} | {row[4]}")
        return str(xlsx_path)
    finally:
        # 只清理本基准创建的内容，绝不整目录删除
        if space_id:
            async with factory() as session:
                await session.execute(text("DELETE FROM document_chunks WHERE space_id = :i"), {"i": space_id})
                await session.execute(text("DELETE FROM documents WHERE space_id = :i"), {"i": space_id})
                await session.execute(text("DELETE FROM knowledge_spaces WHERE id = :i"), {"i": space_id})
                await session.commit()
        for fp in created_files:
            try:
                os.remove(fp)
            except OSError:
                pass
        await engine.dispose()
        print("\n[cleaned] 已清理基准语料与文件")


if __name__ == "__main__":
    t0 = time.time()
    path = asyncio.run(main())
    print(f"\n耗时 {time.time() - t0:.0f}s | 报告: {path}")
